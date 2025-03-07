from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from openpyxl import load_workbook
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import TimeoutException
import pandas as pd
import time
from time import sleep
import tabula
import requests
import fitz
import re
#import jpype
import pdfplumber
import io

####################################################################################


#Skoda
def getSkodaPrices(SkodaModel, SkodaLink):
    
    driver.get(SkodaLink)

    try:
        accept_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@id="onetrust-accept-btn-handler"]'))
        )
        accept_button.click()
    except:
        None

    KonkretKonfigurator_element = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, f"//span[normalize-space()='{SkodaModel}']"))
    )
    KonkretKonfigurator_element.click()
    sleep(3)

    try:
        price_element = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, "//span[contains(@class, 'secondary ng-star-inserted')]"))
        )
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{clean_price}"
        return formatted_price
    except StaleElementReferenceException:
        price_element = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, "//span[contains(@class, 'secondary ng-star-inserted')]"))
        )
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{clean_price}"
        return formatted_price


# Hyundai
def getHyundaiprices(HyundaiModel, HyundaiLink):

    df = tabula.read_pdf(HyundaiLink, pages='all', multiple_tables=True)[0]

    first_space_index = HyundaiModel.find(' ')
    ModellNévElsőSzóközUtán = HyundaiModel[first_space_index+1:]
    target_row = ModellNévElsőSzóközUtán

    try:
            price = df[df.iloc[:, 0] == target_row].iloc[0, 3]
            clean_price = ''.join(filter(str.isdigit, price))
            formatted_price = f"{clean_price}"
    except:
            print("Hibás")
    return formatted_price



# Kia
def getKiaPrices(KiaModel, KiaLink):
    driver.get(KiaLink)
    
    formatted_price = None

    try:
        # Várjunk a gomb betöltésére maximum 10 másodpercig
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )

        # Kattintsunk a gombra
        button.click()

    except:
        pass

    try:
        # Megtaláljuk az adott modellt
        KonkretKonfigurator_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, f"//h3[normalize-space()='{KiaModel}']"))
        )

        # Megkeressük a modellhez tartozó árat
        price_element = KonkretKonfigurator_element.find_element(By.XPATH, "following-sibling::span[contains(@class, 'price')]")
        price = price_element.text

        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{int(clean_price)}"

    except Exception as e:
        print(f"Error finding price for {KiaModel}: {e}")

   
    return formatted_price

#Audi
def getAudiPrices(AudiModel, AudiLink):
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(AudiLink)

    try:
        # Sütik elfogadása
        accept_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@id="onetrust-accept-btn-handler"]'))
        )
        accept_button.click()

        KonkretKonfigurator_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, f'//div[@data-cy-modelclass="{AudiModel}"]'))
        )
        KonkretKonfigurator_element.click()
        time.sleep(3)  # Idő adása az oldalbetöltésre

        konkret = driver.find_element(By.XPATH, f"//span[contains(text(), '{AudiModel}')][@class='group-name-inner']")
        konkret.click()
        time.sleep(4)

        # Az ár kinyerése
        price_element = driver.find_element(By.XPATH, '//span[@class="from"]')
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{int(clean_price)}"

        #print(f"Ár a {audi_model_name} modellhez: {formatted_price}")

        # Visszalépés az előző oldalra
        driver.back()
        time.sleep(3)

    except Exception as e:
        print(f"Hiba történt az Audi modellekkel: {e}")
        formatted_price = None  # Alapértelmezett érték, ha a kivétel elkapódik

    return formatted_price

#Peugeout
def getPeugeotPrices(PeugeotLink):
    formatted_price = 0

    driver.get(PeugeotLink)

    accept_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "_psaihm_id_accept_all_btn"))
        )
    # Elfogadás gombra kattintás
    accept_button.click()
    sleep(3)
    KonkretKonfigurator_element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, f"(//span[@class='q-label'][normalize-space()='KONFIGURÁTOR'])[6]"))
            )
    KonkretKonfigurator_element.click()
    NavigaloGomb_element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, f"//button[@id='cta-slider-next-selection-bar']//i[contains(@role,'presentation')]"))
            )
    NavigaloGomb_element.click()
    sleep(3)
    KonkretModell_element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, f"/html/body/div[1]/div/div[2]/section/div/div[4]/div[1]/div/div/div/div/div/div/div/div/div/div[1]/div[1]/main/section/div[1]/div[1]/div[1]/div/div/div[5]/div/div/div"))
            )
    KonkretModell_element.click()
    sleep(3)
    price_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "(//span[@class='cash-price brand-p'])[1]"))
            )
    price_element = price_element.text
    clean_price = ''.join(filter(str.isdigit, price_element))
    formatted_price = f"{clean_price}"
    return formatted_price
    

#Suzuki
def getSuzukiPrices(SuzukiModel, SuzukiLink):
    driver.get(SuzukiLink)

    try:
        cookie_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'MINDEN SÜTI ENGEDÉLYEZÉSE')]"))
        )
        cookie_button.click()
    except Exception as e:
        print(f"Exception occurred: {e}")

    try:
        models = driver.find_elements(By.XPATH, f"//div[h2[contains(text(), '{SuzukiModel}')]]")
        price_element = models[1].find_element(By.XPATH, ".//following::strong[1]")
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{int(clean_price)}"
        print(f"Price of {SuzukiModel}: {formatted_price}")
    except Exception as e:
        print(f"Error fetching price for {SuzukiModel}: {e}")
        formatted_price = None

    
    return formatted_price


# Lexus
def getLexusPrices(LexusModel, LexusLink):

    driver.get(LexusLink)

    first_space_index = LexusModel.find(' ')
    ModellNévElsőSzóközig = LexusModel[:first_space_index]

    time.sleep(5)
    KonfigurátorElemek = driver.find_elements(By.XPATH, '//a[@data-gt-action="click-cta"]')

    for KonfigurátorElem in KonfigurátorElemek:
        if ModellNévElsőSzóközig in KonfigurátorElem.get_attribute('data-gt-label'):
            KeresettModellKonfigurátora = KonfigurátorElem
            break

    KeresettModellKonfigurátora.click()
    time.sleep(10)

    # Keresse meg az összes olyan elemet, ahol a class tartalmazza az "l-beta-text" szöveget, mert ez az elem veszi fel a modellek konkrét nevét # pl a "https://www.lexus.hu/new-cars/ux/build" oldalon 3 db ilyen elem van: UX 300e, UX 250h AWD, UX 250h FWD
    # MINDEN EGYES LEXUS MODELL ESETÉBEN UGYANEZ A HELYEZ, BÁRMELYIK MODELLNÉL MŰKÖDIK EZ A DINAMIKUS KÓD (AHOGY A TELJES FÜGGVÉNY IS)
    ModellekElemei = driver.find_elements(By.XPATH, '//h2[contains(@class, "l-beta-text")]')

    KeresettModellNév = LexusModel

    # Szűrje az elemeket a keresett modell teljes neve alapján
    for EgyModellEleme in ModellekElemei:
        if KeresettModellNév in EgyModellEleme.text:
            KeresettModellEleme = EgyModellEleme
   
    NagySzülő_div = KeresettModellEleme.find_elements(By.XPATH, './../..')
    Ár_div = NagySzülő_div[0].find_elements(By.XPATH, './/div[contains(@class, "CashPrice__Text")]')
    Ár = Ár_div[0].text

    return Ár

#Cupra
def getCupraPrices(CupraModel, CupraLink):
    driver.get(CupraLink)
    árLine = 0

    try:
        # Várjunk a gomb betöltésére maximum 10 másodpercig
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )

        # Kattintsunk a gombra
        button.click()

    except:
        None
    try:
         # Az árak kinyerése
                price_element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, f"(//span[normalize-space()='{CupraModel}'])"))
                )
                price_element = price_element.find_element(By.XPATH, "..")
                price_element = price_element.find_element(By.XPATH, "..")
                price_element_splitted=price_element.text.split("\n")
                for line in price_element_splitted:
                    if "Ft-tól" in line:
                        árLine = line
                price = árLine
                clean_price = ''.join(filter(str.isdigit, price))
                formatted_price = f"{int(clean_price)}"
    except Exception as e:
        print(f"Hiba történt: {e}")
        formatted_price = None

                
    return formatted_price

# BMW
def getBMWPrices(BMWModel, BMWLink):
    price = None
    formatted_price = None
    clean_price = None
    
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(BMWLink)
    
    extracted_chars = []

    after_bmw = BMWModel.split('BMW ')[-1]
    first_char = after_bmw[0]
    if first_char == 'i':
        extracted_chars.append('BMWi')
    else:
        extracted_chars.append(first_char)
    
    extracted_chars = []

    after_bmw = BMWModel.split('BMW ')[-1]
    first_char = after_bmw[0]
    if first_char == 'i':
            extracted_chars.append('BMWi')
    else:
            extracted_chars.append(first_char)
    sleep(2)
    try:
        cookie_banner = driver.execute_script('''
        return document.querySelector('epaas-consent-drawer-shell').shadowRoot.querySelector('button[class="accept-button button-primary"]');
    ''')
        cookie_banner.click()
    except Exception as e:
        print("A süti elfogadó gomb nem található vagy nem kattintható")
        
    sleep(2)

    for index, char in enumerate(extracted_chars):
        try:
            modellvalasztasbmw = driver.find_element(By.XPATH, f"//a[@class='cmp-modelcard__con-link' and @title='{BMWModel}']")
            modellvalasztasbmw.click()
            sleep(5)  
            
            
            price = driver.execute_script('''
                let elements = document.querySelector("body > con-app").shadowRoot.querySelector("router-slot > con-configure").shadowRoot.querySelector("con-sales-footer").shadowRoot.querySelector("#salesFooterWrapper > div > div.sales-footer-sections > div.sales-footer-sections-top > div.sales-footer-prices-section > div.sales-footer-total-price-wrapper > button > span");
                return elements && elements.textContent;
            ''')
            
            clean_price = ''.join(filter(str.isdigit, price))
            formatted_price = f"{int(clean_price)}"
            
            driver.back()  
            sleep(3)  
            
        except Exception as e:
            print(f"Error with {BMWModel}: {e}")
    time.sleep(2)
    
    return formatted_price


# Volkswagen
def getVolkswagenPrices(VolkswagenModel, VolkswagenLink):
    driver = webdriver.Chrome()
    driver.maximize_window() 
    driver.get(VolkswagenLink)
      



    try:
        accept_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        )
        accept_button.click()
    except:
        pass 
    try: 
        chosen_model = driver.find_element(By.XPATH, f"//span[normalize-space()='{VolkswagenModel}']")
        chosen_model = chosen_model.find_element(By.XPATH, "..").find_element(By.XPATH, "..").find_element(By.XPATH, "..").find_element(By.CLASS_NAME, "image-link")
        chosen_model.click()
        sleep(3)
        
        price_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "(//span[@class='secondary ng-star-inserted'])[1]"))
        )
        
        price = price_element.text
        clean_price = ''.join(filter(str.isdigit, price))
        formatted_price = f"{clean_price}"
        
        #print(f"Ár a {volkswagen_model_name} modellhez: {formatted_price}")
       
        driver.quit()  
    except Exception as e:
        print(f"Error finding price for {VolkswagenModel}: {e}")

    return formatted_price



# Toyota
def GetToyotaPrices(ToyotaModel, ToyotaLink):
    formatted_price = 0

    # Lowercase azért, mert az urlekben csak ilyen betűk vannak, és csak ilyen urlekkel tudjuk megnyitni a weboldalakat
    ToyotaModel = ToyotaModel.lower()

    # Számoljuk meg, hogy szóköz van a beolvasott modellnévben
    SzóközSzám = ToyotaModel.count(' ')

    first_space_index = ToyotaModel.find(' ')
    second_space_index = ToyotaModel.find(' ', first_space_index + 1)

    ModellNévElsőSzóközig = ToyotaModel[:first_space_index].lower()
    ModellnévMásodikSzóközig = ToyotaModel[:second_space_index].lower()

    # MINDEN LEHETSÉGES MODELLNÉV ESET DINAMIKUSAN LEKEZELVE:

    if SzóközSzám == 0: #Például Excelből beolvasott modellnév: Yaris, weboldal: https://www.toyota.hu/modellek/yaris
        ToyotaTeljesUrl = ToyotaLink+"/"+ToyotaModel.lower()
        print(ToyotaTeljesUrl, "0szóköz")
        driver.get(ToyotaTeljesUrl)


    if SzóközSzám == 1: #Például Excelből beolvasott modellnév: Prius Plug-in, weboldal: https://www.toyota.hu/modellek/prius-plugin
        ToyotaTeljesUrl = ToyotaLink+"/"+ToyotaModel.replace('-', "").replace(' ', '-')
        print(ToyotaTeljesUrl, "1szóköz teljesszöveg")
        driver.get(ToyotaTeljesUrl)

        # A Toyota hivatalos oldala úgy van beállítva, hogy ha nem létező urlt próbálunk megnyitni (pl. https://www.toyota.hu/modellek/camry-hybrid),
        # akkor automatikusan átirányít minket (a drivert is természetesen) az Árlisták oldalra. Tehát ha a driver jelenlegi url-je(driver.current_url)
        # nem egyezik azzal az url-lel amellyel az előbb megnyitottuk (driver.get(ToyotaTeljesUrl), akkor biztos, hogy nem létező url-t próbáltunk megnyitni a driverrel
        
        if driver.current_url != ToyotaTeljesUrl: #Például Excelből beolvasott modellnév: Camry Hybrid, weboldal: https://www.toyota.hu/modellek/camry  
            ToyotaTeljesUrl = ToyotaLink+"/"+ModellNévElsőSzóközig
            driver.get(ToyotaTeljesUrl)

    if SzóközSzám >= 2: #Például Excelből beolvasott modellnév: Proace City Verso EV, weboldal: https://www.toyota.hu/modellek/proace-city-verso-ev
        ToyotaTeljesUrl = ToyotaLink+"/"+ToyotaModel.replace('-', "").replace(' ', '-')
        print(ToyotaTeljesUrl, "többszóköz teljesszöveg")
        driver.get(ToyotaTeljesUrl)

        if driver.current_url != ToyotaTeljesUrl: #Például Excelből beolvasott modellnév: Corolla Sedan 1.8 Hybrid, weboldal: https://www.toyota.hu/modellek/corolla-sedan
            ToyotaTeljesUrl = ToyotaLink+"/"+ModellnévMásodikSzóközig.replace('-', "").replace(' ', '-') 
            print(ToyotaTeljesUrl, "többszóköz másodikszóközig")
            driver.get(ToyotaTeljesUrl)

        if driver.current_url != ToyotaTeljesUrl: #Például Excelből beolvasott modellnév: C-HR 1.8 Hybrid, weboldal: https://www.toyota.hu/modellek/c-hr 
            ToyotaTeljesUrl = ToyotaLink+"/"+ModellNévElsőSzóközig 
            driver.get(ToyotaTeljesUrl)

    # Sütik elfogadása
    sleep(2)
    accept_button = WebDriverWait(driver, 2).until(
    EC.element_to_be_clickable((By.XPATH, '//button[@id="onetrust-accept-btn-handler"]')))
    accept_button.click()
    
    # Generalizálható rendszert szeretnénk fejleszteni (hardcoding elkerülése):
    # Az alábbi megoldási módszer ebben a leghatékonyabb, amely a 29 db hivatalosan elérhető Toyota modell közül 5 db modell kivételével bármelyik
    # árát képes visszaadni csupán az input excelben a lekérdezni kívánt új Toyota modell sorának hozzáadásával
    # a 29-ből az 5 db modellnél amelyiknél nem működik, ott többnyire azért nem működik, mert ezek nagy része új = csak előfoglalható (nem elérhető még) = csak előzetes árlista pdf-ek, melyek html és pdf struktúrája nagyon eltér a többi 24 db ugyanolyantól
    # az érintett 5 db modell: Új Yaris, ÚJ Toyota C-HR, ÚJ Land Cruiser, GR Yaris, Yaris VAN

    # A többi 24 db modellnél:
    # Ezek közül számos modell hivatalos weboldalán nem található ár (pl. van "Indulóár (Tartalmazza az ÁFA-t)" - https://www.toyota.hu/modellek/corolla-sedan VS nincs ár -https://www.toyota.hu/modellek/gr86,),
    # azonban a modell weboldalakon található hivatalos Árlista pdf-ekben MIND a 24 db modell esetében megtalálható az ár (pl. https://pdf.sites.toyota.hu/arlista_toyota_c-hr_23.pdf)
    # ezt a pdf url-t minden oldalon annak az elemnek a href attribútum értéke tartalmazza, amely data-gt-action attribútuma = "pricelistdownload"

    # Keresés az elemre a 'data-gt-action="pricelistdownload"' attribútum alapján, mert bármelyik Toyota modell weboldalán ez az element tartalmazza az árlista pdf linkjét
    target_element = driver.find_element(By.XPATH, '//*[@data-gt-action="pricelistdownload"]')

    # Az elem href attribútum értékének kinyerése, azaz az Árlista pdf linkjének kinyerése
    pdf_link = target_element.get_attribute('href')

    response = requests.get(pdf_link)
    response.raise_for_status()

    pdf_file = io.BytesIO(response.content)

    # Megnyitjuk a PDF-et pdfplumber segítségével
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            break

    # Szétbontjuk a pdf-ről kinyert szöveget sorokra
    lines = text.splitlines()

    # Megkeressük azt a sort, amely "tól"-t tartalmaz:
    # BÁRMELYIK hivatalos Toyota Árlista pdf-ben a "Ft-tól"-t tartalmazó "sorban" található a modell ára, példa: https://pdf.sites.toyota.hu/arlista_corolla_sedan_2024.pdf,
    # Azonban a "-" karakter keresésével gyakran adódik probléma (The character U+2011 "‑" could be confused with the ASCII character U+002d "-", which is more common in source code.),
    # ezért csak a "tól" szövegrészletet keressük
    for line in lines:
        if "tól" in line:
            árLine = line
            break

    # Ebből a sorból kivesszük a csak számokat és szóközöket tartalmazó szövegrészeket
    numbers_text = re.sub(r'[^\d\s]', '', árLine)

    # Ahol kettő vagy több szóköz van ebben a szövegben, ott felbontjuk (splitteljük) a szöveget
    # Azért kettő vagy több szóköz, mert ahol csak 1 szóköz van a számjegyek között, az csak az erzescsoportok közti tagolást jelenti 1-1 számnál, pl: "10 260 000  700 000" ( a pdf struktúrában a szövegkinyerésnél ez a 2 szám egy sorban van)
    numbers_str = re.split(r'\s{2,}', numbers_text)

    # Az egyes számok átalakítása egész számokká
    numbers = [int(n.replace(' ', '')) for n in numbers_str]

    # A két szám közül a nagyobbik kiválasztása, 
    # Mivel a kisebbik szám mindig, minden egyes Toyota modell esetében biztosan az árkedvezményt fogja jelenti, míg a nagyobbik szám a modell árát, egy példa: "https://pdf.sites.toyota.hu/arlista_corolla_sedan_2024.pdf"
    formatted_price = max(numbers)

    return formatted_price
  

# Volvo
def GetVolvoPrices(VolvoModel, VolvoLink):


    def extract_text_after_prefix(text, prefix="Kezdőár"):
        # Ellenőrizzük, hogy a szöveg a megadott előtaggal kezdődik-e
        if text.startswith(prefix):
            # Visszaadjuk a szöveget az előtag után, eltávolítva a felesleges részeket
            cleaned_text = text[len(prefix):].strip().replace(" Ft", "").replace(" ", "")
            return int(cleaned_text) if cleaned_text.isdigit() else cleaned_text

        # Ha nincs az előtag, a teljes szöveget tisztítjuk és számmá alakítjuk
        cleaned_text = text.replace(" Ft", "").replace(" ", "")
        return int(cleaned_text) if cleaned_text.isdigit() else cleaned_text


    def get_volvo_price_selenium(url):
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

        driver.get(url)
        driver.maximize_window()

       
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Elfogadás')]"))
            ).click()
        except Exception as e:
            print("Cookie acceptance button not found or another error occurred:", e)

        price = None
        try:
            WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "div.ac > small"))
            )
            price_element = driver.find_element(By.CSS_SELECTOR, "div.ac > small")

            price = extract_text_after_prefix(price_element.text)
            
        except TimeoutException:
            print("Price element not found within the given time.")

        driver.quit()

        return price

    def extract_text_up_to_first_space(text):
        # Megkeressük az első szóközt
        first_space_index = text.find(' ')
        
        # Ha nincs szóköz, akkor a teljes szöveget visszaadjuk
        if first_space_index == -1:
            return text

        # Visszaadjuk a szöveget az első szóközig
        return text[:first_space_index]

    # Teszteljük egy példamondattal

    konkrét_modell_név_urlbe = extract_text_up_to_first_space(VolvoModel)
    konkrét_modell_tipus = df["Meghajtás"].iloc[index,]

    if konkrét_modell_tipus == 'PHEV':
        konkrét_modell_tipus_urlbe = "-hybrid"

    if konkrét_modell_tipus == 'HEV':
        konkrét_modell_tipus_urlbe = "-hybrid"
    
    if konkrét_modell_tipus == 'BEV':
        konkrét_modell_tipus_urlbe = "-electric"

    if konkrét_modell_tipus == 'MHEV':
        konkrét_modell_tipus_urlbe = ""

    volvo_modell_url = VolvoLink+"/cars/"+konkrét_modell_név_urlbe+konkrét_modell_tipus_urlbe
    price = get_volvo_price_selenium(volvo_modell_url)
    print(price, volvo_modell_url)
    
    return price


# MG
def getMGPrices(MGModel, MGLink):
        MGModelUrlbe = MGModel.replace(" ", "_")
  
        driver.get(MGLink+"?model="+MGModelUrlbe)
        # '?model=' azért kell bele, mert egy MG modell konfigurátor linkje minden esetben, minden modellnél a következőképpen néz ki (az 'MG_EHS' rész helyett mindig az aktuális modell név):
        # "https://mggaal.hu/konfigurator?model=MG_EHS", ahol a '?model=' (statikus, nem változó) részen kívül mindent (konfigurátor linkje, aktuális modell neve) az input excel fájlból olvasunk be

        # Sütik elfogadása
        try:
            accept_button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.ID, "cookie-bar-button"))
            )
            accept_button.click()
        except:
            None

        ÁrElementek = WebDriverWait(driver, 5).until(
            EC.presence_of_all_elements_located((By.XPATH, "//span[contains(text(), 'Ft')]"))
        )

        # A legtöbb MG modell esetében 2 ár van feltüntetve a konfigurátor weboldalán, ekkor az egyik MINDIG, BÁRMELYIK MODELL ESETÉBEN át van húzva (elavult ár), míg a másik nincs áthúzva (valós ár)
        # Ennél a résznél a nem áthúzott árat kapjuk vissza
        for ÁrElement in ÁrElementek:
            szülőElement = ÁrElement.find_element(By.XPATH, '..') # A számot (árat) tartalmazó element szülő span-jének a style attribútumában van beállítva az áthúzás mint formázás
            if 'line-through' not in szülőElement.get_attribute('style'): # Ha nincs áthúzva az ár, akkor a számot (árat) tartalmazó element szülő span-jének nincs megadva a style attribútuma (üres string) BÁRMELYIK MODELL ESETÉBEN
                ValósÁr = ÁrElement.text
            
        clean_price = ''.join(filter(str.isdigit, ValósÁr))
        formatted_price = f"{clean_price}"

    
        return formatted_price


####################################################################################


driver = webdriver.Chrome()
driver.maximize_window()

current_datetime = datetime.now()
new_column_name = f'{current_datetime}'

df = pd.read_excel('SAP_arak_lekerdezese.xlsx', header=0)

modellek_neve = df['Modell neve (valós)']
márkák_neve = df['Márka']
modellek_ára = []
type(modellek_neve)

pricelist = []

for index, konkrét_márka_név in enumerate(márkák_neve):
    current_datetime = datetime.now()
    new_column_name = f'{current_datetime}'

df = pd.read_excel('SAP_arak_lekerdezese.xlsx', header=0)

modellek_neve = df['Modell neve (valós)']
márkák_neve = df['Márka']
modellek_ára = []
type(modellek_neve)

pricelist = []

for index, konkrét_márka_név in enumerate(márkák_neve):
   
    
    if konkrét_márka_név == 'Audi':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        AudiLink = df["Link"].iloc[index,]
        modellek_ára.append(getAudiPrices(konkrét_modell_név, AudiLink))
    elif konkrét_márka_név == 'Peugeot':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        PeugeotLink = df["Link"].iloc[index,]
        modellek_ára.append(getPeugeotPrices(PeugeotLink))                     
    elif konkrét_márka_név == 'BMW':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        BMWLink = df["Link"].iloc[index,]
        modellek_ára.append(getBMWPrices(konkrét_modell_név, BMWLink))   
    elif konkrét_márka_név == 'Kia':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        KiaLink = df["Link"].iloc[index,]
        modellek_ára.append(getKiaPrices(konkrét_modell_név, KiaLink))
    elif konkrét_márka_név == 'Suzuki':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        SuzukiLink = df["Link"].iloc[index,]
        modellek_ára.append(getSuzukiPrices(konkrét_modell_név, SuzukiLink))
    elif konkrét_márka_név == 'Cupra':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        CupraLink = df["Link"].iloc[index,]
        modellek_ára.append(getCupraPrices(konkrét_modell_név, CupraLink))    
    elif konkrét_márka_név == 'Volkswagen':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        VolkswagenLink = df["Link"].iloc[index,]
        modellek_ára.append(getVolkswagenPrices(konkrét_modell_név, VolkswagenLink))
    elif konkrét_márka_név == 'MG':
        konkrét_modell_név = df['Modell teljes neve (valós)'].iloc[index,]
        MGLink = df["Link"].iloc[index,]
        modellek_ára.append(getMGPrices(konkrét_modell_név, MGLink))
    elif konkrét_márka_név == 'Hyundai':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        HyundaiLink = df["Link"].iloc[index,]
        modellek_ára.append(getHyundaiprices(konkrét_modell_név, HyundaiLink))
    elif konkrét_márka_név == 'Toyota':
        konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
        ToyotaLink = df["Link"].iloc[index,]
        modellek_ára.append(GetToyotaPrices(konkrét_modell_név, ToyotaLink))
    elif konkrét_márka_név == 'Skoda':
            konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
            SkodaLink = df["Link"].iloc[index,]
            modellek_ára.append(getSkodaPrices(konkrét_modell_név, SkodaLink))
    elif konkrét_márka_név == 'Lexus':
           konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
           LexusLink = df["Link"].iloc[index,]
           modellek_ára.append(getLexusPrices(konkrét_modell_név, LexusLink))
    elif konkrét_márka_név == 'Volvo':
           konkrét_modell_név = df['Modell neve (valós)'].iloc[index,]
           VolvoLink = df["Link"].iloc[index,]
           modellek_ára.append(GetVolvoPrices(konkrét_modell_név, VolvoLink))


else:
     
     print("Baj")

print(modellek_ára)

df_new = pd.DataFrame({f'{current_datetime}': modellek_ára})
wb = load_workbook('SAP_arak_lekerdezese.xlsx')
    
ws = wb['Sheet1']


# Keresse meg az első szabad oszlopot
szabad_oszlop = 'I'
while ws[szabad_oszlop + '1'].value is not None:
    szabad_oszlop = chr(ord(szabad_oszlop) + 1)  # Lépés a következő oszlopra



for index, row in df_new.iterrows():
            cell = f'{szabad_oszlop}%d'  % (index + 2)
            ws[cell] = row[0]

# Az oszlop celláinak formázása, az első sor kihagyásával
for row in range(2, ws.max_row + 1):  # Kezdve a 2. sortól (azaz a G2 cellától)
    cell = ws[f'{szabad_oszlop}{row}']
    if cell.value is not None:  # Ellenőrzés, hogy a cella értéke nem None
        try:
 
            cell.value = int(cell.value)
            cell.number_format = '#,##0 Ft'
        except ValueError:
            # Ha a cella értéke nem alakítható át számmá, hagyja figyelmen kívül
            pass
    else:
        # Itt dönthetsz, hogy mit szeretnél tenni None érték esetén
        print("A cella értéke None")
        cell.value = ''  # állítsuk be 0-ra az értéket

ws[f'{szabad_oszlop}1'] = current_datetime

# Változtatások mentése
wb.save('SAP_arak_lekerdezese.xlsx') 
wb.close()